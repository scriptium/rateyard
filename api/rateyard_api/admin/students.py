import base64
from io import BytesIO
import secrets
import string
import tempfile

from flask import Blueprint, request, abort, jsonify, send_file
from slugify import slugify
import openpyxl

from . import admin_token_required, db

@admin_token_required
def create_students():
    if not request.is_json:
        abort(400, "Expected json")
        print("Expected json", flush=True)
    if type(request.json) != list:
        abort(400, "Expected array of students")
        print("Expected array of students", flush=True)

    database = db.get_db()
    cursor = database.cursor()
    students_data_errors = db.check_students_data(request.json, True)
    was_error = False
    if students_data_errors == {}:
        for student in request.json:
            try:
                cursor.execute('''
                    INSERT INTO students (
                        username,
                        full_name,
                        email,
                        password_hash,
                        class_id
                    )
                    VALUES (
                        %s, %s, %s,
                        crypt(%s, gen_salt('md5')),
                        %s
                    );
                    ''', (
                    student["username"],
                    student["full_name"],
                    student.get("email"),
                    student["password"],
                    student["class_id"]
                ))
            except Exception:
                was_error = True
                
        database.commit()
        if was_error:
            return jsonify({'result': 'Not all students have created'})
        else:
            return jsonify({
                "result": "OK",
            })
    else:
        return jsonify(students_data_errors), 400

@admin_token_required
def delete_students():
    if not request.is_json:
        abort(400, "Expected json")
    if type(request.json) != list:
        abort(400, "Expected array of students id")
    exec_str = '''
            DELETE FROM students
            WHERE False
            '''
    exec_args = []
    for student_id in request.json:
        print(student_id)
        exec_str += (" OR  id=%s")
        exec_args.append(student_id)
    database = db.get_db()
    cursor = database.cursor()
    cursor.execute(exec_str, exec_args)
    database.commit()
    return jsonify({
        "result": "OK"
    }), 200

@admin_token_required
def edit_students():
    print(request.json, flush=True)
    if not request.is_json:
        abort(400, "Expected json")
        print("Expected json", flush=True)
    if type(request.json) != list:
        abort(400, "Expected array of students id")
        print("Expected array of students id", flush=True)

    student_data_errors = db.check_students_data(request.json, False)

    for student_index in range(len(request.json)):
        '''
        Additional error code for students:
        5: wrong student id
        '''
        if not "id" in request.json[student_index].keys():
            student_data_errors[student_index].append(6)

    if student_data_errors == {}:
        for student in request.json:
            db.edit_student(int(student['id']), student) 
        return jsonify(result="ok")

    else:
        return jsonify(student_data_errors), 400

@admin_token_required
def get_students():
    exec_str = '''
    SELECT students.id, students.username, students.full_name,
    students.email, students.class_id, classes.class_name
    FROM students
    LEFT OUTER JOIN classes ON students.class_id=classes.id
    '''
    exec_args = []

    if request.is_json:
        exec_str += " WHERE False"
        for student_id in request.json:
            print('id', student_id, flush=True)
            if type(student_id) != int:
                abort(400)
            exec_args.append(student_id)
            exec_str += " OR students.id=%s"
            
    exec_str += " ORDER BY students.full_name"
    cursor = db.get_db().cursor()
    cursor.execute(exec_str, exec_args)
    exec_result = cursor.fetchall()

    result = [{
        "id": data[0],
        "username": data[1],
        "full_name": data[2],
        "email": data[3],
        "class": {
            "id": data[4],
            "name": data[5],
            "type": "ClassShort"
        },
        "type": "Student"
    } for data in exec_result
    ]
    return jsonify(result)

@admin_token_required
def import_from_excel():
    if not (
        request.is_json and
        type(request.json.get('table_base64')) == str
    ):
        abort(400, 'Expected json with base64 string')

    try:
        table_file = BytesIO(
            base64.b64decode(request.json['table_base64'].encode())
        )
        workbook = openpyxl.load_workbook(table_file)
        worksheet = workbook.active
    except Exception:
        abort(400, 'Invalid excel file')
    
    cursor = db.get_db().cursor()
    cursor.execute(r'SELECT id, class_name FROM classes;')
    classes_dict = {}
    for class_ in cursor.fetchall():
        classes_dict[class_[1]] = class_[0]

    cursor.execute(r'SELECT username FROM students;')
    registered_usernames = set((row[0] for row in cursor.fetchall()))
    students_json = {'students': []}
    alphabet = string.ascii_letters + string.digits
    for row_values in worksheet.iter_rows(values_only=True, min_col=2, min_row=2):
        if len(row_values) >= 5 and all((type(row_values[index]) == str and
                len(row_values[index]) > 0 for index in [*range(3)] + [4])):
            username = slugify(row_values[0], separator='')
            if username in registered_usernames:
                username_number = 1
                new_username = username + str(username_number)
                while new_username in registered_usernames:
                    username_number += 1
                    new_username = username + str(username_number)

                username = new_username

            registered_usernames.add(username)
            students_json['students'].append({
                'username': username,
                'full_name': ' '.join((row_values[0], row_values[1], row_values[2])),
                'class': {
                    'id': classes_dict[row_values[4].upper()],
                    'name': row_values[4].upper()
                },
                'password': ''.join(secrets.choice(alphabet) for i in range(5))
            })
    
    if len(students_json['students']) == 0:
        abort(400, 'No valide students found')
    
    passwords_workbook = openpyxl.workbook.Workbook()
    passwords_worksheet = passwords_workbook.active
    passwords_worksheet.append(('ПІБ', 'Логін', 'Клас', 'Пароль'))
    for student in students_json['students']:
        passwords_worksheet.append((
            student['full_name'],
            student['username'],
            student['class']['name'],
            student['password']
        ))

    passwords_file = tempfile.NamedTemporaryFile()
    passwords_workbook.save(passwords_file.name)
    students_json['passwords_table_base64'] = base64.b64encode(passwords_file.read()).decode()

    return jsonify(students_json)